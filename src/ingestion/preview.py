"""Preview locally downloaded Toronto Open Data files under data/raw."""

from __future__ import annotations

import argparse
import csv
import io
import struct
import sys
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

DEFAULT_RAW_DIR = Path(__file__).resolve().parents[2] / "data" / "raw"
DEFAULT_ROWS = 5
TEXT_SUFFIXES = {".csv", ".tsv", ".txt", ".json", ".geojson", ".md", ".prj"}
TABULAR_SUFFIXES = {".csv", ".tsv"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}


@dataclass(frozen=True)
class FilePreview:
    """Structured preview of one local data file (or archive).

    Attributes:
        path: Absolute or relative path that was previewed.
        size_bytes: On-disk file size in bytes.
        kind: High-level type label (``csv``, ``zip``, ``xlsx``, ``dbf``, ...).
        summary: Short human-readable summary line.
        columns: Column / field names when available.
        rows: Sample rows as lists of string values.
        members: Archive member names for ZIP files.
        extras: Optional key/value metadata (sheet names, record counts, etc.).
    """

    path: Path
    size_bytes: int
    kind: str
    summary: str
    columns: list[str] = field(default_factory=list)
    rows: list[list[str]] = field(default_factory=list)
    members: list[str] = field(default_factory=list)
    extras: dict[str, Any] = field(default_factory=dict)


def list_raw_files(raw_dir: Path, dataset: str | None = None) -> list[Path]:
    """List downloadable data files under the raw directory.

    Args:
        raw_dir: Root raw directory (typically ``data/raw``).
        dataset: Optional dataset key subdirectory to restrict to
            (e.g. ``travel_times_bluetooth``).

    Returns:
        Sorted list of file paths, excluding ``.gitkeep`` and ``*.partial``.
    """
    root = raw_dir / dataset if dataset else raw_dir
    if not root.exists():
        return []

    files: list[Path] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if path.name == ".gitkeep" or path.name.endswith(".partial"):
            continue
        files.append(path)
    return files


def _sample_csv_text(
    text_stream: Iterable[str],
    n_rows: int,
) -> tuple[list[str], list[list[str]], int | None]:
    """Read column names and sample rows from a text CSV stream.

    Args:
        text_stream: Iterable of decoded text lines.
        n_rows: Maximum number of data rows to return.

    Returns:
        Tuple of ``(columns, sample_rows, sniffed_dialect_note)`` where the
        third value is unused placeholder reserved for future metadata.
    """
    reader = csv.reader(text_stream)
    try:
        columns = next(reader)
    except StopIteration:
        return [], [], None

    rows: list[list[str]] = []
    for row in reader:
        rows.append(row)
        if len(rows) >= n_rows:
            break
    return columns, rows, None


def preview_csv(path: Path, n_rows: int = DEFAULT_ROWS) -> FilePreview:
    """Preview a standalone CSV/TSV file.

    Args:
        path: Path to the CSV file.
        n_rows: Number of data rows to include in the sample.

    Returns:
        ``FilePreview`` with columns and sample rows.
    """
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        columns, rows, _ = _sample_csv_text(handle, n_rows)
    return FilePreview(
        path=path,
        size_bytes=path.stat().st_size,
        kind="csv",
        summary=f"{len(columns)} columns, showing {len(rows)} row(s)",
        columns=columns,
        rows=rows,
    )


def preview_text(path: Path, max_chars: int = 500) -> FilePreview:
    """Preview a small plain-text file.

    Args:
        path: Path to the text file.
        max_chars: Maximum number of characters to include.

    Returns:
        ``FilePreview`` whose ``summary`` contains the truncated text.
    """
    raw = path.read_text(encoding="utf-8", errors="replace")
    snippet = raw[:max_chars]
    if len(raw) > max_chars:
        snippet += "…"
    return FilePreview(
        path=path,
        size_bytes=path.stat().st_size,
        kind="text",
        summary=snippet.replace("\n", "\\n"),
    )


def _dbf_field_names(data: bytes) -> tuple[list[str], int]:
    """Parse DBF header field names and record count.

    Args:
        data: Raw bytes of a ``.dbf`` file (at least the header).

    Returns:
        Tuple of ``(field_names, record_count)``.
    """
    if len(data) < 32:
        return [], 0
    record_count = struct.unpack("<I", data[4:8])[0]
    header_length = struct.unpack("<H", data[8:10])[0]
    names: list[str] = []
    offset = 32
    while offset + 32 <= len(data) and offset < header_length:
        if data[offset] == 0x0D:
            break
        raw_name = data[offset : offset + 11]
        name = raw_name.split(b"\x00", 1)[0].decode("ascii", errors="replace").strip()
        if name:
            names.append(name)
        offset += 32
    return names, record_count


def preview_zip(path: Path, n_rows: int = DEFAULT_ROWS) -> FilePreview:
    """Preview a ZIP archive, sampling the first tabular member when present.

    Args:
        path: Path to the ZIP file.
        n_rows: Number of CSV rows to sample from the first tabular member.

    Returns:
        ``FilePreview`` including member list and optional CSV sample.
    """
    members: list[str] = []
    columns: list[str] = []
    rows: list[list[str]] = []
    extras: dict[str, Any] = {}

    with zipfile.ZipFile(path) as archive:
        infos = [
            info
            for info in archive.infolist()
            if not info.is_dir() and not Path(info.filename).name.startswith("._")
            and "__MACOSX" not in info.filename
        ]
        members = [f"{info.filename} ({info.file_size} bytes)" for info in infos]

        # Prefer CSV/TSV samples; otherwise expose DBF field names for shapefiles.
        for info in infos:
            suffix = Path(info.filename).suffix.lower()
            if suffix in TABULAR_SUFFIXES:
                with archive.open(info) as binary:
                    text = io.TextIOWrapper(binary, encoding="utf-8", errors="replace", newline="")
                    columns, rows, _ = _sample_csv_text(text, n_rows)
                extras["sampled_member"] = info.filename
                extras["member_bytes"] = info.file_size
                break
        else:
            for info in infos:
                if Path(info.filename).suffix.lower() == ".dbf":
                    with archive.open(info) as binary:
                        data = binary.read(4096)
                    field_names, record_count = _dbf_field_names(data)
                    columns = field_names
                    extras["sampled_member"] = info.filename
                    extras["dbf_records"] = record_count
                    break

    summary_bits = [f"{len(members)} member(s)"]
    if "sampled_member" in extras:
        summary_bits.append(f"sampled {extras['sampled_member']}")
    if columns:
        summary_bits.append(f"{len(columns)} columns/fields")

    return FilePreview(
        path=path,
        size_bytes=path.stat().st_size,
        kind="zip",
        summary=", ".join(summary_bits),
        columns=columns,
        rows=rows,
        members=members,
        extras=extras,
    )


def preview_xlsx(path: Path, n_rows: int = DEFAULT_ROWS) -> FilePreview:
    """Preview an Excel workbook (requires ``openpyxl``).

    Args:
        path: Path to the ``.xlsx`` file.
        n_rows: Number of data rows to sample from the first sheet
            (header is separate).

    Returns:
        ``FilePreview`` with sheet names, columns, and sample rows.

    Raises:
        ImportError: If ``openpyxl`` is not installed.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - exercised when dep missing
        raise ImportError(
            "openpyxl is required to preview Excel files. "
            "Install with: pip install openpyxl"
        ) from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        sheet = workbook[sheet_names[0]]
        extracted = list(sheet.iter_rows(values_only=True, max_row=n_rows + 1))
    finally:
        workbook.close()

    columns = ["" if v is None else str(v) for v in (extracted[0] if extracted else [])]
    rows = [
        ["" if v is None else str(v) for v in row]
        for row in extracted[1:]
    ]
    return FilePreview(
        path=path,
        size_bytes=path.stat().st_size,
        kind="xlsx",
        summary=f"sheets={sheet_names}; showing {len(rows)} row(s) from '{sheet_names[0]}'",
        columns=columns,
        rows=rows,
        extras={"sheets": sheet_names},
    )


def preview_file(path: Path, n_rows: int = DEFAULT_ROWS) -> FilePreview:
    """Dispatch preview logic based on file extension.

    Args:
        path: Local file path.
        n_rows: Sample row count for tabular sources.

    Returns:
        ``FilePreview`` for the file. Unknown types get a size-only summary.
    """
    suffix = path.suffix.lower()
    if suffix == ".zip":
        return preview_zip(path, n_rows=n_rows)
    if suffix in TABULAR_SUFFIXES:
        return preview_csv(path, n_rows=n_rows)
    if suffix in EXCEL_SUFFIXES:
        return preview_xlsx(path, n_rows=n_rows)
    if suffix in TEXT_SUFFIXES:
        return preview_text(path)
    return FilePreview(
        path=path,
        size_bytes=path.stat().st_size,
        kind=suffix.lstrip(".") or "unknown",
        summary="no tabular preview available for this file type",
    )


def format_preview(preview: FilePreview) -> str:
    """Render a ``FilePreview`` as plain text for terminal output.

    Args:
        preview: Structured preview object.

    Returns:
        Multi-line string ready to print.
    """
    lines = [
        f"## {preview.path}",
        f"size: {preview.size_bytes:,} bytes",
        f"kind: {preview.kind}",
        f"summary: {preview.summary}",
    ]
    if preview.members:
        lines.append("members:")
        for member in preview.members:
            lines.append(f"  - {member}")
    if preview.columns:
        lines.append("columns: " + ", ".join(preview.columns))
    if preview.rows:
        lines.append("sample rows:")
        width = max(len(preview.columns), max((len(r) for r in preview.rows), default=0))
        header = preview.columns + [""] * (width - len(preview.columns))
        lines.append("  " + " | ".join(header))
        lines.append("  " + "-+-".join("-" * max(len(c), 1) for c in header))
        for row in preview.rows:
            padded = row + [""] * (width - len(row))
            lines.append("  " + " | ".join(padded))
    if preview.extras:
        interesting = {
            key: value
            for key, value in preview.extras.items()
            if key not in {"sampled_member"}
        }
        if interesting:
            lines.append("extras: " + ", ".join(f"{k}={v}" for k, v in interesting.items()))
    return "\n".join(lines)


def _build_parser() -> argparse.ArgumentParser:
    """Build the CLI argument parser for dataset previews.

    Args:
        None.

    Returns:
        Configured ``argparse.ArgumentParser``.
    """
    parser = argparse.ArgumentParser(
        description="Preview downloaded Toronto Open Data files under data/raw."
    )
    parser.add_argument(
        "--raw-dir",
        type=Path,
        default=DEFAULT_RAW_DIR,
        help="Root raw directory (default: data/raw).",
    )
    parser.add_argument(
        "--dataset",
        help="Dataset key subdirectory to preview (e.g. travel_times_bluetooth).",
    )
    parser.add_argument(
        "--file",
        type=Path,
        help="Preview a single file path instead of scanning a dataset folder.",
    )
    parser.add_argument(
        "--rows",
        type=int,
        default=DEFAULT_ROWS,
        help=f"Number of sample rows for tabular files (default: {DEFAULT_ROWS}).",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    """CLI entrypoint: print previews for raw dataset files.

    Args:
        argv: Optional argument list (without the program name).

    Returns:
        ``0`` on success, ``1`` if no matching files were found.
    """
    args = _build_parser().parse_args(argv)

    if args.file:
        paths = [args.file]
    else:
        paths = list_raw_files(args.raw_dir, args.dataset)

    if not paths:
        target = args.file or (args.raw_dir / args.dataset if args.dataset else args.raw_dir)
        print(f"No files found to preview under: {target}", file=sys.stderr)
        return 1

    blocks = [format_preview(preview_file(path, n_rows=args.rows)) for path in paths]
    print("\n\n".join(blocks))
    return 0


if __name__ == "__main__":
    sys.exit(main())
